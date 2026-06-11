#!/usr/bin/env python3
"""Generate key_lut columns + named-glyph additions for the 2026-06 "Europe +
Americas" minority/sibling batch (Wave 1 — Latin only, no new font):

  eu-ES gl-ES rm-CH cy-GB ga-IE mt-MT lb-LU se-NO   (Europe)
  gn-PY qu-PE ay-BO nv-US nh-MX                       (Americas)

Column types:
  * pure clones    — cell-for-cell copies of an existing column (incl. settings):
                     eu-ES gl-ES <- es-ES (Spanish, Basque/Galician use it)
                     rm-CH <- de-CH (Swiss German QWERTZ, Romansh)
                     lb-LU <- fr-CH (Swiss French, Luxembourg)
                     gn-PY qu-PE ay-BO nh-MX <- es-MX (Latin-American "latam")
  * clone+override — clone a base column then patch a few keys:
                     cy-GB <- en-GB + AltGr ŵ/ŷ (Welsh to-bach)
                     ga-IE <- en-GB + AltGr acute fadas á é í ó ú (xkb ie)
                     mt-MT <- en-GB + ċ ġ ħ ż + €/£ + grave vowels (xkb mt)
  * partial new    — fold base (en-US) + AltGr letters:
                     nv-US Navajo: AltGr ł ą ę į ǫ on L A E I O
                     se-NO Northern Sami: á š ŧ č ž đ ŋ + å ø æ home keys (xkb
                       no(smi)); the moved-out q w y x show as AltGr previews

Codes are protocol-constrained to 2+2 chars. nh-MX (Nahuatl) has no ISO-639-1
code → the pseudo-code "nh" must be appended to PRIVATE_LANGS in the frozen
iso_lang_country.py (all 3 repos) before re-cogging hid_com.c. Every other code
here is standard ISO 639-1 / 3166-1 (no frozen-table change).

⚠ NULL-base trap (see FUTURE_LANGUAGES.md world-batch post-fixes): a key with a
NULL VAR_SMALL makes translate_keycode fall back to en-US *entirely* for that
key, so its VAR_SHIFT/VAR_ALTGR is never shown. Every override key below carries
an explicit base.

Outputs: /tmp/euam_cols.json (for _patch_xlsx.py),
         /tmp/euam_named.json (for _patch_named_glyphs.py)
"""
import json
from openpyxl import load_workbook

XLSX = "lang_lut.xlsx"

ROW = {}
for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
    ROW["KC_" + c] = 2 + i
for i in range(1, 10):
    ROW[f"KC_{i}"] = 27 + i
ROW["KC_0"] = 37
ROW.update({"KC_MINUS": 43, "KC_EQUAL": 44, "KC_LBRC": 45, "KC_RBRC": 46,
            "KC_BACKSLASH": 47, "KC_NONUS_HASH": 48, "KC_SEMICOLON": 49,
            "KC_QUOTE": 50, "KC_GRAVE": 51, "KC_COMMA": 52, "KC_DOT": 53,
            "KC_SLASH": 54, "KC_NONUS_BACKSLASH": 55})

S = lambda v: ["str", v]
N = lambda v: ["num", v]

# Settings (rows 56-61); fr-FR/Latin style with AltGr preview offsets.
# letter.voffset[VAR_ALTGR] = 9 (NOT 13): the tall _SupAndExtA_/_LatinExtB_ glyphs
# (yAdvance 44 vs the base 40) and ogonek/descender letters (ą ę į ǫ, ŷ, q, y)
# draw ~4 px low (oled_preview.py models this as gy += yAdvance-base_yAdvance), so
# 13 clipped them at the 40 px panel bottom. 9 lifts them clear — same fix as the
# world batch. Verify with PolyKybdHost/tools/oled_preview.py --lang <code>.
SET_LATIN = {56: [None, S("HIDE"), None, N(50)], 57: [None, S("HIDE"), None, N(9)],
             58: [None, N(28), None, N(50)],     59: [None, N(0), None, N(13)],
             60: [None, N(28), None, N(50)],     61: [None, N(0), None, N(13)]}

# Override applied to the clone-based AltGr-letter layouts so they inherit the
# clone's hoffset/num/sym settings but use the un-clipped letter AltGr voffset.
ALTGR_V9 = {57: [None, S("HIDE"), None, N(9)]}

# ── new named glyphs (Latin-1 / Ext-A/B / Ext-Additional) ────────────────────
NAMED = [["A_WITH_ACUTE", "C1"], ["A_WITH_ACUTE_SMALL", "E1"],   # Á á (Irish/Sami fada)
         ["LATIN_01EA", "1EA"], ["LATIN_01EB", "1EB"],           # Ǫ ǫ (Navajo)
         ["LATIN_1EBD", "1EBD"], ["LATIN_1EF9", "1EF9"]]         # ẽ ỹ (Guarani nasal)

# ── helper: a letter key [small, SHIFT, CAPS=None, ALTGR] ─────────────────────
def L(small, shift, altgr=None):
    return [S(small), S(shift), None, (S(altgr) if altgr else None)]

# ── clone+override / new-mapping tables ──────────────────────────────────────
# Welsh: UK layout + circumflex (to-bach) ŵ ŷ on AltGr.
CY_GB = {"KC_W": L("w", "W", "LATIN_0175"), "KC_Y": L("y", "Y", "LATIN_0177")}

# Irish: UK layout + acute fadas on AltGr (xkb ie "basic").
GA_IE = {"KC_A": L("a", "A", "A_WITH_ACUTE_SMALL"),
         "KC_E": L("e", "E", "E_WITH_ACUTE_SMALL"),
         "KC_I": L("i", "I", "I_WITH_ACUTE_SMALL"),
         "KC_O": L("o", "O", "O_WITH_ACUTE_SMALL"),
         "KC_U": L("u", "U", "U_WITH_ACUTE_SMALL")}

# Maltese: UK base + the four accented-dot letters ċ ġ ħ ż, €/£ and grave vowels
# (xkb mt "basic"). ċ ġ ħ ż are letters (CAPS=None → capslock shows the capital).
MT_MT = {
    "KC_GRAVE":           [S("LATIN_010B"), S("LATIN_010A"), None, S("GRAVE_ACCENT")],  # ċ Ċ
    "KC_LBRC":            [S("LATIN_0121"), S("LATIN_0120"), None, S("[")],             # ġ Ġ
    "KC_RBRC":            [S("LATIN_0127"), S("LATIN_0126"), None, S("]")],             # ħ Ħ
    "KC_NONUS_BACKSLASH": [S("LATIN_017C"), S("LATIN_017B"), None, S("BACKSLASH")],     # ż Ż
    "KC_2": [N(2), S("QUOTE"),     N(2), S("SUPER_SCRIPT_2")],
    "KC_3": [N(3), S("EURO_SIGN"), N(3), S("POUND_SIGN")],
    "KC_4": [N(4), S("$"),         N(4), S("EURO_SIGN")],
    "KC_QUOTE": [S("'"), S("@"), S("'"), S("^")],
    "KC_A": L("a", "A", "A_WITH_GRAVE_SMALL"),
    "KC_E": L("e", "E", "E_WITH_GRAVE_SMALL"),
    "KC_I": L("i", "I", "I_WITH_GRAVE_SMALL"),
    "KC_O": L("o", "O", "O_WITH_GRAVE_SMALL"),
    "KC_U": L("u", "U", "U_WITH_GRAVE_SMALL"),
}

# Navajo: US base + AltGr ł ą ę į ǫ (the distinctive consonant + nasal vowels).
NV_US = {
    "KC_L": L("l", "L", "LATIN_0142"),   # ł
    "KC_A": L("a", "A", "LATIN_0105"),   # ą
    "KC_E": L("e", "E", "LATIN_0119"),   # ę
    "KC_I": L("i", "I", "LATIN_012F"),   # į
    "KC_O": L("o", "O", "LATIN_01EB"),   # ǫ
}

# Northern Sami: home-row Sami letters with the displaced q/w/y/x as AltGr
# previews; Nordic å ø æ on the [ ; ' keys (xkb no(smi)). Rest folds to en-US.
SE_NO = {
    "KC_Q": L("A_WITH_ACUTE_SMALL", "A_WITH_ACUTE", "q"),  # á Á
    "KC_W": L("LATIN_0161", "LATIN_0160", "w"),            # š Š
    "KC_Y": L("LATIN_0167", "LATIN_0166", "y"),            # ŧ Ŧ
    "KC_X": L("LATIN_010D", "LATIN_010C", "x"),            # č Č
    "KC_NONUS_BACKSLASH": L("LATIN_017E", "LATIN_017D"),   # ž Ž
    "KC_BACKSLASH":       L("LATIN_0111", "LATIN_0110", "'"),  # đ Đ
    "KC_RBRC":            L("LATIN_014B", "LATIN_014A"),   # ŋ Ŋ
    "KC_LBRC":            L("LATIN_00E5", "LATIN_00C5"),   # å Å
    "KC_SEMICOLON":       L("LATIN_00F8", "LATIN_00D8"),   # ø Ø
    "KC_QUOTE":           L("LATIN_00E6", "LATIN_00C6"),   # æ Æ
}

# Guarani: types on the Latin-American (latam) layout — the nasal vowels are made
# with the dead tilde — so it clones es-MX and adds the six nasal vowels as AltGr
# previews (ã ẽ ĩ õ ũ ỹ) so the keycaps actually show Guarani's distinctive set.
GN_PY = {
    "KC_A": L("a", "A", "LATIN_00E3"),   # ã
    "KC_E": L("e", "E", "LATIN_1EBD"),   # ẽ
    "KC_I": L("i", "I", "LATIN_0129"),   # ĩ
    "KC_O": L("o", "O", "LATIN_00F5"),   # õ
    "KC_U": L("u", "U", "LATIN_0169"),   # ũ
    "KC_Y": L("y", "Y", "LATIN_1EF9"),   # ỹ
}

# ── clone helper (mirrors _gen_world_cols.clone_col) ─────────────────────────
def clone_col(sheet, langs, src):
    base = langs[src]
    out = {}
    for r in range(2, 62):
        cells = []
        for k in range(4):
            v = sheet.cell(row=r, column=base + k).value
            if v is None:
                cells.append(None)
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                cells.append(N(int(v)))
            else:
                cells.append(S(str(v)))
        if any(c is not None for c in cells):
            out[r] = cells
    return out

wb = load_workbook(XLSX, data_only=True, read_only=True)
sh = wb["key_lut"]
langs = {}
i = 0
while True:
    v = sh.cell(row=1, column=2 + i * 4).value
    if not v:
        break
    langs[v] = 2 + i * 4
    i += 1

def overlay(base_cols, override):
    """clone dict + per-key override (override rows replace cloned rows)."""
    out = dict(base_cols)
    for k, quad in override.items():
        out[ROW[k]] = quad
    return out

def partial(override, settings):
    cols = {ROW[k]: v for k, v in override.items()}
    cols.update(settings)
    return cols

SPEC = {
    # pure clones
    "eu-ES": clone_col(sh, langs, "es-ES"),
    "gl-ES": clone_col(sh, langs, "es-ES"),
    "rm-CH": clone_col(sh, langs, "de-CH"),
    "lb-LU": clone_col(sh, langs, "fr-CH"),
    "gn-PY": overlay(clone_col(sh, langs, "es-MX"), GN_PY),
    "qu-PE": clone_col(sh, langs, "es-MX"),
    "ay-BO": clone_col(sh, langs, "es-MX"),
    "nh-MX": clone_col(sh, langs, "es-MX"),
    # clone + override
    "cy-GB": overlay(clone_col(sh, langs, "en-GB"), CY_GB),
    "ga-IE": overlay(clone_col(sh, langs, "en-GB"), GA_IE),
    "mt-MT": overlay(clone_col(sh, langs, "en-GB"), MT_MT),
    # partial new mappings (fold base)
    "nv-US": partial(NV_US, SET_LATIN),
    "se-NO": partial(SE_NO, SET_LATIN),
}
ORDER = ["eu-ES", "gl-ES", "rm-CH", "cy-GB", "ga-IE", "mt-MT", "lb-LU", "se-NO",
         "gn-PY", "qu-PE", "ay-BO", "nv-US", "nh-MX"]
assert set(ORDER) == set(SPEC)

# Clone-based AltGr-letter layouts inherit a v=13 letter voffset from their clone
# source; switch just that to the un-clipped 9 (nv-US/se-NO already get it via
# SET_LATIN). cy/ga/mt have accented AltGr letters; gn adds the nasal vowels.
for _L in ("cy-GB", "ga-IE", "mt-MT", "gn-PY"):
    SPEC[_L][57] = ALTGR_V9[57]

spec_json = {n: {str(r): v for r, v in SPEC[n].items()} for n in ORDER}
json.dump(spec_json, open("/tmp/euam_cols.json", "w"), indent=1)
json.dump(NAMED, open("/tmp/euam_named.json", "w"), indent=1)
print("order:", ",".join(ORDER))
print(f"{len(ORDER)} languages, {len(NAMED)} new named glyphs")
