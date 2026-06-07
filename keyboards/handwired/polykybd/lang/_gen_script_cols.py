#!/usr/bin/env python3
"""Generate key_lut columns + named-glyph additions + fonts.yaml matra sequences
for the Indic/Thai batch (th-TH, bn-IN, te-IN, ta-IN).

Bengali/Telugu/Tamil are InScript clones of hi-IN: every Devanagari codepoint is
shifted into the target Unicode block (ISCII-aligned: +0x80 / +0x300 / +0x280)
and validated against the target font's cmap; codepoints with no glyph in the
target script are dropped (left NULL -> English fallback). Thai is the Kedmanee
layout transcribed from xkb symbols/th. Combining marks (matras / tone marks)
are emitted as dotted-circle composites at per-script PUA ranges, mirroring the
DEVA_DC_* mechanism.

Outputs:
  /tmp/script_cols.json        key_lut columns for _patch_xlsx.py
  /tmp/named_glyphs_add.json   [[name, hexcode], ...] contiguous append to sheet1
  /tmp/fonts_yaml_add.yaml     the ALL_FONTS entries to splice into fonts.yaml
"""
import json, re, os, tempfile
from pathlib import Path
from fontTools.ttLib import TTFont

FONTS = str(Path(__file__).resolve().parents[1] / "fonts") + "/"   # repo-relative
OUT = tempfile.gettempdir()
KEYSYMDEF = os.environ.get("KEYSYMDEF_H", "/usr/include/X11/keysymdef.h")
def cmap_of(p):
    return set(TTFont(FONTS+p).getBestCmap().keys())

CMAP = {
 "bn": cmap_of("noto-sans-bengali/NotoSansBengali-Regular.ttf"),
 "te": cmap_of("noto-sans-telugu/NotoSansTelugu-Regular.ttf"),
 "ta": cmap_of("noto-sans-tamil/NotoSansTamil-Regular.ttf"),
 "th": cmap_of("noto-sans-thai/NotoSansThai-Regular.ttf"),
}

# ---- QMK keycode -> key_lut row -------------------------------------------
ROW={}; L="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
for i,c in enumerate(L): ROW["KC_"+c]=2+i
for i in range(1,10): ROW[f"KC_{i}"]=27+i
ROW["KC_0"]=37
ROW.update({"KC_MINUS":43,"KC_EQUAL":44,"KC_LBRC":45,"KC_RBRC":46,
            "KC_BACKSLASH":47,"KC_NONUS_HASH":48,"KC_SEMICOLON":49,
            "KC_QUOTE":50,"KC_GRAVE":51,"KC_COMMA":52,"KC_DOT":53,
            "KC_SLASH":54,"KC_NONUS_BACKSLASH":55})
SETTINGS_ROWS=[56,57,58,59,60,61]

named_add=[]          # (name, hexcode) contiguous
seen_names=set()
def add_named(name, cp):
    if name not in seen_names:
        named_add.append([name, f"{cp:04X}"]); seen_names.add(name)

cols_out={}
fonts_yaml=[]

# ============================================================================
# Indic InScript clones (shift hi-IN)
# ============================================================================
from openpyxl import load_workbook
wb=load_workbook(FONTS+"../lang/lang_lut.xlsx", data_only=True, read_only=True)
sh=wb['key_lut']
hi=None; i=0
while True:
    v=sh.cell(row=1,column=2+i*4).value
    if v=="hi-IN": hi=2+i*4; break
    i+=1
labels={r:sh.cell(row=r,column=1).value for r in range(2,62)}
hi_cells={r:[sh.cell(row=r,column=hi+k).value for k in range(4)] for r in range(2,62)}

def parse_deva(tok):
    """return ('base',cp) | ('dc',cp) | ('lit',s) | None for a hi-IN cell token"""
    if tok in (None,''): return None
    m=re.match(r'DEVANAGARI_([0-9A-Fa-f]{3,4})$', str(tok))
    if m: return ('base', int(m.group(1),16))
    m=re.match(r'DEVA_DC_([0-9A-Fa-f]{3,4})$', str(tok))
    if m: return ('dc', int(m.group(1),16))
    return ('lit', str(tok))

INDIC=[("bn-IN","bn",0x80,"BENGALI","BENG_DC",0xE120),
       ("te-IN","te",0x300,"TELUGU","TELU_DC",0xE140),
       ("ta-IN","ta",0x280,"TAMIL","TAML_DC",0xE160)]

def make_conv(cm, off, PFX, DCPFX, dc_marks):
    """Indic shift+cmap converter; explicit params avoid loop-variable capture (B023)."""
    def conv(tok):
        p=parse_deva(tok)
        if p is None: return None
        kind=p[0]
        if kind=='lit': return ("str", p[1])
        cp=p[1]+off
        if cp not in cm: return None          # no glyph in target script
        if kind=='base':
            add_named(f"{PFX}_{cp:04X}", cp); return ("str", f"{PFX}_{cp:04X}")
        else:                                  # combining mark -> composite
            dc_marks.add(cp); return ("str", f"{DCPFX}_{cp:04X}")
    return conv

for langname, key, off, PFX, DCPFX, pua_base in INDIC:
    cm=CMAP[key]
    dc_marks=set()                 # target mark codepoints used -> composite
    conv=make_conv(cm, off, PFX, DCPFX, dc_marks)
    colmap={}
    for r in range(2,56):
        quad=hi_cells.get(r,[None]*4)
        cells=[conv(quad[0]), conv(quad[1]), conv(quad[2]), conv(quad[3])]
        if any(c is not None for c in cells):
            colmap[r]=cells
    # settings: copy hi-IN's offset block verbatim
    for r in SETTINGS_ROWS:
        q=hi_cells[r]
        cells=[]
        for v in q:
            if v in (None,''): cells.append(None)
            elif isinstance(v,(int,float)): cells.append(["num",int(v)])
            else: cells.append(["str",str(v)])
        colmap[r]=cells
    cols_out[langname]=colmap
    # PUA assignment for the marks, sorted by codepoint
    seq_parts=[]
    for j,mark in enumerate(sorted(dc_marks)):
        pua=pua_base+j
        add_named(f"{DCPFX}_{mark:04X}", pua)
        seq_parts.append(f"25CC {mark:04X}")
    if seq_parts:
        fonts_yaml.append({
            "lang":langname,"key":key,"block":(0x900+off,0x900+off+0x7f),
            "pua_base":pua_base,"seq":", ".join(seq_parts)})

# ============================================================================
# Thai (Kedmanee, from xkb symbols/th)
# ============================================================================
# keysym name -> codepoint, parsed from keysymdef.h (Thai_*), plus ASCII helpers
THAI_KS={}
if not os.path.exists(KEYSYMDEF):
    raise SystemExit(f"keysymdef.h not found at {KEYSYMDEF} - install libx11-dev or set KEYSYMDEF_H")
for line in open(KEYSYMDEF):
    m=re.search(r'#define XK_(Thai_\w+)\s+0x[0-9a-fA-F]+\s+/\*\s*U\+([0-9A-Fa-f]{4})', line)
    if m: THAI_KS[m.group(1)]=int(m.group(2),16)

ASCII={'percent':'%','plus':'+','slash':'/','minus':'-','quotedbl':'QUOTE',
       'comma':'COMMA','period':'.','question':'?','parenleft':'(',
       'parenright':')','underscore':'_'}

# QMK key -> (base_keysym, shift_keysym) from the xkb dump
TH={
 "KC_GRAVE":("underscore","percent"),
 "KC_1":("Thai_lakkhangyao","plus"),"KC_2":("slash","Thai_leknung"),
 "KC_3":("minus","Thai_leksong"),"KC_4":("Thai_phosamphao","Thai_leksam"),
 "KC_5":("Thai_thothung","Thai_leksi"),"KC_6":("Thai_sarau","Thai_sarauu"),
 "KC_7":("Thai_saraue","Thai_baht"),"KC_8":("Thai_khokhwai","Thai_lekha"),
 "KC_9":("Thai_totao","Thai_lekhok"),"KC_0":("Thai_chochan","Thai_lekchet"),
 "KC_MINUS":("Thai_khokhai","Thai_lekpaet"),"KC_EQUAL":("Thai_chochang","Thai_lekkao"),
 "KC_Q":("Thai_maiyamok","Thai_leksun"),"KC_W":("Thai_saraaimaimalai","quotedbl"),
 "KC_E":("Thai_saraam","Thai_dochada"),"KC_R":("Thai_phophan","Thai_thonangmontho"),
 "KC_T":("Thai_saraa","Thai_thothong"),"KC_Y":("Thai_maihanakat","Thai_nikhahit"),
 "KC_U":("Thai_saraii","Thai_maitri"),"KC_I":("Thai_rorua","Thai_nonen"),
 "KC_O":("Thai_nonu","Thai_paiyannoi"),"KC_P":("Thai_yoyak","Thai_yoying"),
 "KC_LBRC":("Thai_bobaimai","Thai_thothan"),"KC_RBRC":("Thai_loling","comma"),
 "KC_A":("Thai_fofan","Thai_ru"),"KC_S":("Thai_hohip","Thai_khorakhang"),
 "KC_D":("Thai_kokai","Thai_topatak"),"KC_F":("Thai_dodek","Thai_sarao"),
 "KC_G":("Thai_sarae","Thai_chochoe"),"KC_H":("Thai_maitho","Thai_maitaikhu"),
 "KC_J":("Thai_maiek","Thai_maichattawa"),"KC_K":("Thai_saraaa","Thai_sorusi"),
 "KC_L":("Thai_sosua","Thai_sosala"),"KC_SEMICOLON":("Thai_wowaen","Thai_soso"),
 "KC_QUOTE":("Thai_ngongu","period"),
 "KC_BACKSLASH":("Thai_khokhuat","Thai_khokhon"),
 "KC_NONUS_HASH":("Thai_khokhuat","Thai_khokhon"),
 "KC_Z":("Thai_phophung","parenleft"),"KC_X":("Thai_popla","parenright"),
 "KC_C":("Thai_saraae","Thai_choching"),"KC_V":("Thai_oang","Thai_honokhuk"),
 "KC_B":("Thai_sarai","Thai_phinthu"),"KC_N":("Thai_sarauee","Thai_thanthakhat"),
 "KC_M":("Thai_thothahan","question"),"KC_COMMA":("Thai_moma","Thai_thophuthao"),
 "KC_DOT":("Thai_saraaimaimuan","Thai_lochula"),"KC_SLASH":("Thai_fofa","Thai_lu"),
}
THAI_COMBINING=set([0x0E31]+list(range(0x0E34,0x0E3B))+list(range(0x0E47,0x0E4F)))
th_cm=CMAP["th"]; th_marks=set()
def th_conv(ks):
    if ks is None: return None
    if ks in ASCII:
        v=ASCII[ks]; return ("str", v)
    cp=THAI_KS.get(ks)
    if cp is None or cp not in th_cm: return None
    if cp in THAI_COMBINING:
        th_marks.add(cp); return ("str", f"THAI_DC_{cp:04X}")
    add_named(f"THAI_{cp:04X}", cp); return ("str", f"THAI_{cp:04X}")
thcol={}
for k,(b,s) in TH.items():
    r=ROW[k]; cb=th_conv(b); cs=th_conv(s)
    thcol[r]=[cb, cs, cb, None]        # CAPS=base, no altgr
# Thai settings: copy hi-IN offset block (same font size class)
for r in SETTINGS_ROWS:
    q=hi_cells[r]; cells=[]
    for v in q:
        if v in (None,''): cells.append(None)
        elif isinstance(v,(int,float)): cells.append(["num",int(v)])
        else: cells.append(["str",str(v)])
    thcol[r]=cells
cols_out["th-TH"]=thcol
th_seq=[]
for j,mark in enumerate(sorted(th_marks)):
    pua=0xE180+j; add_named(f"THAI_DC_{mark:04X}", pua); th_seq.append(f"25CC {mark:04X}")
if th_seq:
    fonts_yaml.append({"lang":"th-TH","key":"th","block":(0xE00,0xE7F),
                       "pua_base":0xE180,"seq":", ".join(th_seq)})

json.dump(cols_out, open(os.path.join(OUT,"script_cols.json"),"w"), indent=1)
json.dump(named_add, open(os.path.join(OUT,"named_glyphs_add.json"),"w"), indent=1)
json.dump(fonts_yaml, open(os.path.join(OUT,"fonts_yaml_add.json"),"w"), indent=1)
print(f"langs: {list(cols_out)}")
for ln,cm in cols_out.items():
    nonset=[r for r in cm if r<56 and any(c is not None for c in cm[r])]
    print(f"  {ln}: {len(nonset)} mapped keys")
print(f"named glyph additions: {len(named_add)}  (first {named_add[0]} .. last {named_add[-1]})")
for fy in fonts_yaml:
    print(f"  fonts.yaml {fy['lang']}: block {hex(fy['block'][0])}-{hex(fy['block'][1])}, {len(fy['seq'].split(','))} matra composites @ PUA {hex(fy['pua_base'])}")
