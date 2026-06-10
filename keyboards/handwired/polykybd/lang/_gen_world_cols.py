#!/usr/bin/env python3
"""Generate key_lut columns + named-glyph additions for the 2026-06 "world"
batch — the Oceania/Africa candidates from CLAUDE.md plus the per-tab
computer-user picks (see lang/FUTURE_LANGUAGES.md):

  en-AU en-NZ mi-NZ sm-WS fj-FJ tl-PH hw-US en-ZA af-ZA ar-EG sw-KE am-ET
  yo-NG en-NG ar-MA ar-IQ ku-IQ ms-MY uz-UZ en-CA es-AR en-PG ty-PF

Column types:
  * folds        — empty key cells (en-US fallback) + id-ID-style settings:
                   en-AU en-NZ fj-FJ tl-PH en-ZA sw-KE ms-MY en-CA en-PG
  * clones       — cell-for-cell copies of an existing column:
                   ar-EG/ar-MA/ar-IQ <- ar-SA, es-AR <- es-MX, ty-PF <- fr-FR
  * new mappings — mi-NZ (xkb nz/mao), sm-WS/hw-US (Polynesian macrons+okina),
                   af-ZA (xkb za AltGr accents, Latin-1/Ext-A subset),
                   yo-NG/en-NG (xkb ng), uz-UZ (xkb uz/latin),
                   am-ET (xkb et/olpc, Ethiopic), ku-IQ (xkb ir/ku_ara, Sorani)

Language codes are protocol-constrained to 2+2 chars ("hw-US" stands in for
ISO-639-2 "haw", "ku-IQ" for Sorani ckb, "en-PG" covers PNG incl. Tok Pisin).

Outputs: /tmp/world_cols.json (for _patch_xlsx.py),
         /tmp/world_named.json (for _patch_named_glyphs.py)
"""
import json, sys
from openpyxl import load_workbook

XLSX = "lang_lut.xlsx"

ROW = {}
for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
    ROW["KC_" + c] = 2 + i
for i in range(1, 10):
    ROW[f"KC_{i}"] = 27 + i
ROW["KC_0"] = 37
ROW.update({"KC_MINUS": 43, "KC_EQUAL": 44, "KC_LBRC": 45, "KC_RBRC": 46,
            "KC_BACKSLASH": 47, "KC_NONUS_HASH": 48, "KC_SEMICOLON": 49,
            "KC_QUOTE": 50, "KC_GRAVE": 51, "KC_COMMA": 52, "KC_DOT": 53,
            "KC_SLASH": 54, "KC_NONUS_BACKSLASH": 55})

S = lambda v: ["str", v]
N = lambda v: ["num", v]

# Settings blocks (rows 56-61), one 4-tuple [SMALL,SHIFT,CAPS,ALTGR] per row.
SET_FOLD   = {56: [None, S("HIDE"), None, None], 57: [None, S("HIDE"), None, None],
              58: [None, N(35), None, None],     59: [None, N(0), None, None],
              60: [None, N(35), None, None],     61: [None, N(0), None, None]}
SET_LATIN  = {56: [None, S("HIDE"), None, N(50)], 57: [None, S("HIDE"), None, N(13)],
              58: [None, N(28), None, N(50)],     59: [None, N(0), None, N(13)],
              60: [None, N(28), None, N(50)],     61: [None, N(0), None, N(13)]}
SET_ARABIC = {56: [None, N(42), None, N(52)], 57: [None, N(0), None, N(13)],
              58: [None, N(27), None, N(52)], 59: [None, N(0), None, N(13)],
              60: [None, N(42), None, N(52)], 61: [None, N(0), None, N(13)]}

# ── new named glyphs ─────────────────────────────────────────────────────────
ETHIOPIC_CPS = [0x1200, 0x1208, 0x1210, 0x1218, 0x1220, 0x1228, 0x1230, 0x1238,
                0x1240, 0x1250, 0x1260, 0x1268, 0x1270, 0x1278, 0x1280, 0x1290,
                0x1298, 0x12A0, 0x12A1, 0x12A2, 0x12A4, 0x12A6, 0x12A8, 0x12B8,
                0x12C8, 0x12D0, 0x12D8, 0x12E0, 0x12E8, 0x12F0, 0x12F8, 0x1300,
                0x1308, 0x1318, 0x1320, 0x1328, 0x1330, 0x1338, 0x1340, 0x1348,
                0x1350, 0x1361, 0x1362, 0x1363, 0x1364, 0x1365, 0x1366, 0x1367,
                0x1369, 0x136A, 0x136B, 0x136C, 0x136D, 0x136E, 0x136F, 0x1370,
                0x1371, 0x137B]
NAMED = [["OKINA", "2BB"],            # ʻ Hawaiian okina / Samoan koma liliu / Uzbek tutuq
         ["NAIRA_SIGN", "20A6"],      # ₦ (en-NG / yo-NG, xkb ng Shift+4)
         ["LATIN_1E62", "1E62"], ["LATIN_1E63", "1E63"],   # Ṣ ṣ (Yoruba)
         ["LATIN_1EB8", "1EB8"], ["LATIN_1EB9", "1EB9"],   # Ẹ ẹ (Yoruba)
         ["LATIN_1ECC", "1ECC"], ["LATIN_1ECD", "1ECD"],   # Ọ ọ (Yoruba)
         # Sorani Kurdish extras (beyond the fa-IR Perso-Arabic set)
         ["ARABIC_REH_V", "695"],     # ڕ
         ["ARABIC_VEH", "6A4"],       # ڤ
         ["ARABIC_LAM_V", "6B5"],     # ڵ
         ["ARABIC_OE", "6C6"],        # ۆ
         ["ARABIC_YEH_V", "6CE"],     # ێ
         ["ARABIC_AE", "6D5"],        # ە
         ] + [[f"ETHIOPIC_{cp:04X}", f"{cp:04X}"] for cp in ETHIOPIC_CPS]

E = lambda cp: S(f"ETHIOPIC_{cp:04X}")

# ── per-language key tables ──────────────────────────────────────────────────
MACRONS = {  # AltGr macron vowels (xkb nz/mao pattern), small form shown
    "KC_A": [None, None, None, S("LATIN_0101")],
    "KC_E": [None, None, None, S("LATIN_0113")],
    "KC_I": [None, None, None, S("LATIN_012B")],
    "KC_O": [None, None, None, S("LATIN_014D")],
    "KC_U": [None, None, None, S("LATIN_016B")],
}

MI_NZ = dict(MACRONS)

SM_WS = dict(MACRONS)
SM_WS["KC_QUOTE"] = [None, None, None, S("OKINA")]

HW_US = dict(MACRONS)
# Hawaiian layout: okina on the apostrophe key, ASCII apostrophe via AltGr.
HW_US["KC_QUOTE"] = [S("OKINA"), S("QUOTE"), S("OKINA"), S("'")]

AF_ZA = {  # xkb za(basic) AltGr accents, Latin-1/Ext-A subset (Afrikaans set)
    "KC_A": [None, None, None, S("UMLAUT_A_SMALL")],   # ä
    "KC_E": [None, None, None, S("LATIN_00EB")],       # ë
    "KC_R": [None, None, None, S("LATIN_00EA")],       # ê
    "KC_Y": [None, None, None, S("LATIN_00FB")],       # û
    "KC_U": [None, None, None, S("UMLAUT_U_SMALL")],   # ü
    "KC_I": [None, None, None, S("LATIN_00EF")],       # ï
    "KC_O": [None, None, None, S("UMLAUT_O_SMALL")],   # ö
    "KC_P": [None, None, None, S("LATIN_00F4")],       # ô
    "KC_S": [None, None, None, S("LATIN_0161")],       # š
    "KC_2": [None, None, None, S("YEN_SIGN")],
    "KC_3": [None, None, None, S("POUND_SIGN")],
    "KC_5": [None, None, None, S("EURO_SIGN")],
    "KC_6": [None, None, None, S("^")],
    "KC_GRAVE": [None, None, None, S("GRAVE_ACCENT")],
    "KC_SEMICOLON": [None, None, None, S("DIAERESIS")],
    "KC_QUOTE": [None, None, None, S("ACUTE_ACCENT")],
    "KC_DOT": [None, None, None, S("MOD_DOT_ACCENT")],
}

NAIRA = {"KC_4": [None, S("NAIRA_SIGN"), None, S("$")]}  # xkb ng: Shift+4 = ₦

YO_NG = {  # xkb ng(yoruba): ẹ/ọ/ṣ replace q/x/v (Latin letter via AltGr)
    "KC_Q": [S("LATIN_1EB9"), S("LATIN_1EB8"), None, S("q")],
    "KC_X": [S("LATIN_1ECD"), S("LATIN_1ECC"), None, S("x")],
    "KC_V": [S("LATIN_1E63"), S("LATIN_1E62"), None, S("v")],
    **NAIRA,
}

UZ_UZ = {"KC_QUOTE": [S("OKINA"), S("QUOTE"), S("OKINA"), None]}  # xkb uz(latin)

AM_ET = {  # xkb et(olpc); vowel keys show the independent vowels the dead keys produce
    "KC_A": [E(0x12A0), None, E(0x12A0), None],
    "KC_B": [E(0x1260), None, E(0x1260), None],
    "KC_C": [E(0x1278), None, E(0x1278), None],
    "KC_D": [E(0x12F0), E(0x12F8), E(0x12F0), None],
    "KC_E": [E(0x12A4), None, E(0x12A4), None],
    "KC_F": [E(0x1348), None, E(0x1348), None],
    "KC_G": [E(0x1308), E(0x1318), E(0x1308), None],
    "KC_H": [E(0x1200), E(0x1210), E(0x1200), None],
    "KC_I": [E(0x12A2), None, E(0x12A2), None],
    "KC_J": [E(0x1300), None, E(0x1300), None],
    "KC_K": [E(0x12A8), E(0x12B8), E(0x12A8), None],
    "KC_L": [E(0x1208), None, E(0x1208), None],
    "KC_M": [E(0x1218), None, E(0x1218), None],
    "KC_N": [E(0x1290), E(0x1298), E(0x1290), None],
    "KC_O": [E(0x12A6), None, E(0x12A6), None],
    "KC_P": [E(0x1350), E(0x1330), E(0x1350), None],
    "KC_Q": [E(0x1240), E(0x1250), E(0x1240), None],
    "KC_R": [E(0x1228), None, E(0x1228), None],
    "KC_S": [E(0x1230), E(0x1220), E(0x1230), None],
    "KC_T": [E(0x1270), E(0x1320), E(0x1270), None],
    "KC_U": [E(0x12A1), None, E(0x12A1), None],
    "KC_V": [E(0x1238), E(0x1268), E(0x1238), None],
    "KC_W": [E(0x12C8), None, E(0x12C8), None],
    "KC_X": [E(0x12A0), E(0x12D0), E(0x12A0), None],
    "KC_Y": [E(0x12E8), None, E(0x12E8), None],
    "KC_Z": [E(0x12D8), E(0x12E0), E(0x12D8), None],
    # Ge'ez numerals as AltGr previews (ar-SA digit-row pattern)
    "KC_1": [N(1), S("!"), N(1), E(0x1369)],
    "KC_2": [N(2), S("@"), N(2), E(0x136A)],
    "KC_3": [N(3), S("#"), N(3), E(0x136B)],
    "KC_4": [N(4), S("$"), N(4), E(0x136C)],
    "KC_5": [N(5), S("%"), N(5), E(0x136D)],
    "KC_6": [N(6), S("^"), N(6), E(0x136E)],
    "KC_7": [N(7), S("&"), N(7), E(0x136F)],
    "KC_8": [N(8), S("*"), N(8), E(0x1370)],
    "KC_9": [N(9), S("("), N(9), E(0x1371)],
    "KC_0": [S("ZERO"), S(")"), S("ZERO"), E(0x137B)],
    # Ethiopic punctuation
    "KC_SEMICOLON": [E(0x1362), E(0x1361), E(0x1362), None],  # ። ፡
    "KC_QUOTE": [E(0x1366), E(0x1365), E(0x1366), None],      # ፦ ፥
    "KC_COMMA": [E(0x1363), None, E(0x1363), None],           # ፣
    "KC_DOT": [E(0x1364), None, E(0x1364), None],             # ፤
    "KC_SLASH": [E(0x1367), S("?"), E(0x1367), None],         # ፧
}

KU_IQ = {  # xkb ir(ku_ara) — Sorani Kurdish (Arabic script)
    "KC_A": [S("ARABIC_ALEF"), S("ARABIC_ALEF_MADDA_A"), S("ARABIC_ALEF"), None],
    "KC_B": [S("ARABIC_BEH"), S("ARABIC_ALEF_MAKSURA"), S("ARABIC_BEH"), None],
    "KC_C": [S("ARABIC_JEEM"), S("ARABIC_TCHEH"), S("ARABIC_JEEM"), None],
    "KC_D": [S("ARABIC_DAL"), S("ARABIC_THAL"), S("ARABIC_DAL"), None],
    "KC_E": [S("ARABIC_AE"), S("ARABIC_HEH"), S("ARABIC_AE"), None],
    "KC_F": [S("ARABIC_FEH"), S("ARABIC_ALEF_HAMZA_B"), S("ARABIC_FEH"), None],
    "KC_G": [S("ARABIC_GAF"), S("ARABIC_GHAIN"), S("ARABIC_GAF"), None],
    "KC_H": [S("ARABIC_HEH"), None, S("ARABIC_HEH"), None],
    "KC_I": [S("ARABIC_HAH"), S("ARABIC_AIN"), S("ARABIC_HAH"), None],
    "KC_J": [S("ARABIC_JEH"), S("ARABIC_ALEF_HAMZA_A"), S("ARABIC_JEH"), None],
    "KC_K": [S("ARABIC_KEHEH"), S("ARABIC_KAF"), S("ARABIC_KEHEH"), None],
    "KC_L": [S("ARABIC_LAM"), S("ARABIC_LAM_V"), S("ARABIC_LAM"), None],
    "KC_M": [S("ARABIC_MEEM"), S("ARABIC_TATWEEL"), S("ARABIC_MEEM"), None],
    "KC_N": [S("ARABIC_NOON"), S("ARABIC_TEH_MARBUTA"), S("ARABIC_NOON"), None],
    "KC_O": [S("ARABIC_OE"), S("ARABIC_WAW_HAMZA_A"), S("ARABIC_OE"), None],
    "KC_P": [S("ARABIC_PEH"), S("ARABIC_THEH"), S("ARABIC_PEH"), None],
    "KC_Q": [S("ARABIC_QAF"), None, S("ARABIC_QAF"), None],
    "KC_R": [S("ARABIC_REH"), S("ARABIC_REH_V"), S("ARABIC_REH"), None],
    "KC_S": [S("ARABIC_SEEN"), S("ARABIC_SHEEN"), S("ARABIC_SEEN"), None],
    "KC_T": [S("ARABIC_TEH"), S("ARABIC_TAH"), S("ARABIC_TEH"), None],
    "KC_U": [S("ARABIC_YEH_HAMZA_A"), S("ARABIC_HAMZA"), S("ARABIC_YEH_HAMZA_A"), None],
    "KC_V": [S("ARABIC_VEH"), S("ARABIC_ZAH"), S("ARABIC_VEH"), None],
    "KC_W": [S("ARABIC_WAW"), None, S("ARABIC_WAW"), None],
    "KC_X": [S("ARABIC_KHAH"), S("ARABIC_SAD"), S("ARABIC_KHAH"), None],
    "KC_Y": [S("ARABIC_FARSI_YEH"), S("ARABIC_YEH_V"), S("ARABIC_FARSI_YEH"), None],
    "KC_Z": [S("ARABIC_ZAIN"), S("ARABIC_DAD"), S("ARABIC_ZAIN"), None],
    "KC_1": [N(1), S("!"), N(1), S("ARABIC_INDIC_1")],
    "KC_2": [N(2), S("@"), N(2), S("ARABIC_INDIC_2")],
    "KC_3": [N(3), S("#"), N(3), S("ARABIC_INDIC_3")],
    "KC_4": [N(4), S("$"), N(4), S("ARABIC_INDIC_4")],
    "KC_5": [N(5), S("%"), N(5), S("ARABIC_INDIC_5")],
    "KC_6": [N(6), S("^"), N(6), S("ARABIC_INDIC_6")],
    "KC_7": [N(7), S("&"), N(7), S("ARABIC_INDIC_7")],
    "KC_8": [N(8), S("*"), N(8), S("ARABIC_INDIC_8")],
    "KC_9": [N(9), S(")"), N(9), S("ARABIC_INDIC_9")],
    "KC_0": [S("ZERO"), S("("), S("ZERO"), S("ARABIC_INDIC_0")],
    "KC_MINUS": [S("-"), S("ARABIC_TATWEEL"), S("-"), None],
    "KC_LBRC": [S("]"), S("}"), S("]"), None],
    "KC_RBRC": [S("["), S("{"), S("["), None],
    "KC_SEMICOLON": [S("ARABIC_SEMICOLON"), S(":"), S("ARABIC_SEMICOLON"), None],
    "KC_GRAVE": [None, S("DEVISION_SIGN"), None, S("~")],
    "KC_COMMA": [S("ARABIC_COMMA"), S(">"), S("ARABIC_COMMA"), None],
    "KC_DOT": [None, S("<"), None, None],
    "KC_SLASH": [None, S("ARABIC_QMARK"), None, S("?")],
}

# ── clone helper ─────────────────────────────────────────────────────────────
def clone_col(sheet, langs, src):
    base = langs[src]
    out = {}
    for r in range(2, 62):
        cells = []
        for k in range(4):
            v = sheet.cell(row=r, column=base + k).value
            if v is None:
                cells.append(None)
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                cells.append(N(int(v)))
            else:
                cells.append(S(str(v)))
        if any(c is not None for c in cells):
            out[r] = cells
    return out

wb = load_workbook(XLSX, data_only=True, read_only=True)
sh = wb["key_lut"]
langs = {}
i = 0
while True:
    v = sh.cell(row=1, column=2 + i * 4).value
    if not v:
        break
    langs[v] = 2 + i * 4
    i += 1

def build(table, settings):
    cols = {ROW[k]: v for k, v in table.items()}
    cols.update(settings)
    return cols

SPEC = {
    "en-AU": build({}, SET_FOLD),
    "en-NZ": build({}, SET_FOLD),
    "mi-NZ": build(MI_NZ, SET_LATIN),
    "sm-WS": build(SM_WS, SET_LATIN),
    "fj-FJ": build({}, SET_FOLD),
    "tl-PH": build({}, SET_FOLD),
    "hw-US": build(HW_US, SET_LATIN),
    "en-ZA": build({}, SET_FOLD),
    "af-ZA": build(AF_ZA, SET_LATIN),
    "ar-EG": clone_col(sh, langs, "ar-SA"),
    "sw-KE": build({}, SET_FOLD),
    "am-ET": build(AM_ET, SET_ARABIC),
    "yo-NG": build(YO_NG, SET_LATIN),
    "en-NG": build(NAIRA, SET_LATIN),
    "ar-MA": clone_col(sh, langs, "ar-SA"),
    "ar-IQ": clone_col(sh, langs, "ar-SA"),
    "ku-IQ": build(KU_IQ, SET_ARABIC),
    "ms-MY": build({}, SET_FOLD),
    "uz-UZ": build(UZ_UZ, SET_LATIN),
    "en-CA": build({}, SET_FOLD),
    "es-AR": clone_col(sh, langs, "es-MX"),
    "en-PG": build({}, SET_FOLD),
    "ty-PF": clone_col(sh, langs, "fr-FR"),
}
ORDER = list(SPEC.keys())

# JSON keys must be strings
spec_json = {n: {str(r): v for r, v in cols.items()} for n, cols in SPEC.items()}
json.dump(spec_json, open("/tmp/world_cols.json", "w"), indent=1)
json.dump(NAMED, open("/tmp/world_named.json", "w"), indent=1)
print("order:", ",".join(ORDER))
print(f"{len(ORDER)} languages, {len(NAMED)} new named glyphs")
