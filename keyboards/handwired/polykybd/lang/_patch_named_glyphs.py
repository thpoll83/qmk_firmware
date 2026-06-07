#!/usr/bin/env python3
"""Append named-glyph rows to the named_glyphs sheet (sheet1.xml) of lang_lut.xlsx.

Reads /tmp/named_glyphs_add.json = [[name, hexcode], ...]. The named_glyphs.h cog
loop reads col A (name) + col B (hex codepoint) until the first blank row, so the
new rows MUST be contiguous after the last filled row (1104 = DEVA_DC_094D). Rows
1105-1125 exist as data-free blanks; we drop them and append all new rows in
order, mimicking the inlineStr format the existing DEVA_DC_* rows already use.
Only sheet1.xml is rewritten; every other zip entry (incl. the formula caches in
latin_sup_ex / the rest of named_glyphs) is copied byte-for-byte.
"""
import sys, json, re, zipfile, shutil, os, tempfile
from xml.sax.saxutils import escape

if len(sys.argv) < 3:
    sys.exit("usage: _patch_named_glyphs.py <xlsx> <named_glyphs_add.json>")
XLSX=sys.argv[1]
ADD=json.load(open(sys.argv[2]))
# first blank row after the last contiguous filled (name+code) row
from openpyxl import load_workbook
_wb=load_workbook(XLSX,data_only=True,read_only=True); _sh=_wb['named_glyphs']
_r=2
while _sh.cell(row=_r,column=1).value not in (None,'') and _sh.cell(row=_r,column=2).value not in (None,''):
    _r+=1
_wb.close()
START=_r                                     # first blank row after last filled glyph

def rowxml(n, name, code):
    return (f'<row r="{n}" customFormat="false" ht="12.8" hidden="false" '
            f'customHeight="false" outlineLevel="0" collapsed="false">'
            f'<c r="A{n}" t="inlineStr"><is><t xml:space="preserve">{escape(name)}</t></is></c>'
            f'<c r="B{n}" t="inlineStr"><is><t xml:space="preserve">{escape(code)}</t></is></c>'
            f'</row>')

tmp=tempfile.mkdtemp(prefix="ng_work_")
with zipfile.ZipFile(XLSX) as z: z.extractall(tmp)
p=os.path.join(tmp,"xl/worksheets/sheet1.xml")
xml=open(p,encoding="utf-8").read()

# drop pre-existing blank rows >= START
xml=re.sub(r'<row r="(\d+)"[^>]*>.*?</row>',
           lambda m: '' if int(m.group(1))>=START else m.group(0), xml, flags=re.S)

new_rows="".join(rowxml(START+i, n, c) for i,(n,c) in enumerate(ADD))
xml=xml.replace("</sheetData>", new_rows+"</sheetData>")

last=START+len(ADD)-1
xml=re.sub(r'<dimension ref="A1:I\d+"/>', f'<dimension ref="A1:I{last}"/>', xml)
open(p,"w",encoding="utf-8").write(xml)

out=XLSX+".new"
with zipfile.ZipFile(XLSX) as zin, zipfile.ZipFile(out,"w",zipfile.ZIP_DEFLATED) as zout:
    for it in zin.infolist():
        data=open(p,"rb").read() if it.filename=="xl/worksheets/sheet1.xml" else zin.read(it.filename)
        zout.writestr(it,data)
shutil.move(out,XLSX)
print(f"appended {len(ADD)} named glyphs (rows {START}..{last})")
