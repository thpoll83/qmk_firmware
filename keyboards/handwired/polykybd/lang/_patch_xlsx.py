#!/usr/bin/env python3
"""Surgically append language columns to key_lut (sheet2.xml) inside lang_lut.xlsx.

Reads a JSON spec produced by the column generators:
  { lang_name: { row_number(str): [c0,c1,c2,c3] } }   cell = None|["num",n]|["str",s]
Appends one 4-column group per language *after* the current last language,
writes ONLY xl/worksheets/sheet2.xml (every other zip entry is copied byte-for-
byte so formula caches in sheet1/sheet3/sharedStrings stay intact), updates
<dimension>, and rewrites the .xlsx in place.

Optional LCID map adds the Windows locale id in the 4th header sub-column
(purely informational; cog never reads it).
"""
import sys, json, re, zipfile, shutil, os
from xml.sax.saxutils import escape

XLSX = sys.argv[1]
SPEC = json.load(open(sys.argv[2]))
ORDER = sys.argv[3].split(",")            # explicit lang order to append
LCID = {"en-GB":2057,"es-MX":2058,"de-CH":2055,"fr-BE":2060,"fr-CA":3084,
        "th-TH":1054,"bn-IN":1093,"te-IN":1098,"ta-IN":1097}

def colname(n):                            # 1-based index -> A1 letters
    s=""
    while n:
        n,r=divmod(n-1,26); s=chr(65+r)+s
    return s

def num_cell(ref,v):  return f'<c r="{ref}"><v>{v}</v></c>'
def str_cell(ref,v):  return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escape(v)}</t></is></c>'

# locate current language count (base col 2 + i*4 until blank header)
from openpyxl import load_workbook
wb = load_workbook(XLSX, data_only=True, read_only=True)
sh = wb['key_lut']
n=0
while sh.cell(row=1, column=2+n*4).value: n+=1
wb.close()
print(f"existing languages: {n} (next base col = {2+n*4})")

# build per-row appended cell XML
# rows that get cells: 1 (header) + every row present in any lang spec
appended = {}                              # row -> list of (col, xml)
for i,name in enumerate(ORDER):
    base = 2 + (n+i)*4
    # header row 1: name at base col, optional LCID at base+3
    appended.setdefault(1,[]).append((base, str_cell(f"{colname(base)}1", name)))
    if name in LCID:
        appended[1].append((base+3, num_cell(f"{colname(base+3)}1", LCID[name])))
    cols = SPEC[name]
    for rs, quad in cols.items():
        r=int(rs)
        for k,cellval in enumerate(quad):
            if cellval is None: continue
            c = base+k
            ref = f"{colname(c)}{r}"
            kind,val = cellval
            xml = num_cell(ref,val) if kind=="num" else str_cell(ref,val)
            appended.setdefault(r,[]).append((c,xml))

newmaxcol = 2 + (n+len(ORDER))*4 - 1       # last col used (incl LCID at base+3)
print(f"new max col index = {newmaxcol} ({colname(newmaxcol)})")

# read sheet2.xml
tmp="/tmp/_xlsx_work"; shutil.rmtree(tmp,ignore_errors=True); os.makedirs(tmp)
with zipfile.ZipFile(XLSX) as z:
    names=z.namelist()
    z.extractall(tmp)
sheet2=os.path.join(tmp,"xl/worksheets/sheet2.xml")
xml=open(sheet2,encoding="utf-8").read()

# inject appended cells before each row's </row>
def inject(m):
    head=m.group(0)
    rn=int(m.group(1))
    if rn not in appended: return head
    extra="".join(x for _,x in sorted(appended[rn], key=lambda t:t[0]))
    return head[:-len("</row>")]+extra+"</row>"
xml=re.sub(r'<row r="(\d+)"[^>]*>.*?</row>', inject, xml, flags=re.S)

# update dimension
xml=re.sub(r'<dimension ref="A1:[A-Z]+(\d+)"/>',
           lambda m: f'<dimension ref="A1:{colname(newmaxcol)}{m.group(1)}"/>', xml)
open(sheet2,"w",encoding="utf-8").write(xml)

# rewrite zip: copy every entry, swapping sheet2
out=XLSX+".new"
with zipfile.ZipFile(XLSX) as zin, zipfile.ZipFile(out,"w",zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data=open(os.path.join(tmp,item.filename),"rb").read() if item.filename=="xl/worksheets/sheet2.xml" else zin.read(item.filename)
        zout.writestr(item, data)
shutil.move(out, XLSX)
print("patched", XLSX)
