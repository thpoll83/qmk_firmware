#!/usr/bin/env python3
"""Append a variation to a letter's row in the latin_sup_ex sheet (sheet3.xml) of
lang_lut.xlsx, in place.  Only sheet3.xml is rewritten; every other zip entry --
and therefore every cached formula result cog reads with data_only=True -- is
copied byte-for-byte.

⚠️ This exists because openpyxl MUST NOT save this workbook: it does not evaluate
formulas, so a round-trip drops every cached <v> and the next run_cog.sh emits
empty tables with no error.  See lang/README.md.

Each letter occupies two rows: the odd row holds the label in column A and the
variation CHARACTERS in B onwards, the row below holds =DEC2HEX(UNICODE(..)) over
each of them -- and it is that cached hex that cog reads.  So appending a
variation means writing both cells, plus the hex the formula would have produced.
"""
import sys, re, zipfile, shutil, os, unicodedata
import xml.dom.minidom as minidom
from xml.sax.saxutils import escape

SHEET = "xl/worksheets/sheet3.xml"

# Variations live in columns B.. ; the per-row Min/Max helpers sit to their RIGHT.
# ⚠️ Those helpers used to be at N/O, i.e. only ten slots past B, and the free-column
# scan below happily walked into them: appending an 11th/12th variation to a row
# overwrote the MIN/MAX formula cells, and the fixup pass then crashed trying to
# int() a hex cached value it had just written there itself.  They now live at S/T
# ⚠️ The helper position is DERIVED, never a literal.  It has been overrun twice:
# first at N/O (ten slots past B), which the free-column scan walked into as soon
# as a row gained an 11th variation; then at S/T, which held only while MAX_SLOTS
# was 16 -- raising it to 64 for the 6-bit widening put the scan straight through
# them again, and the first run wrote a variation into the MIN cell.  Deriving the
# columns from FIRST_COL + MAX_SLOTS makes the collision unrepresentable instead of
# merely further away.  Use _move_helper_columns.py if MAX_SLOTS ever grows again.
#
# MAX_SLOTS is also the hard ceiling of the picker's stored index: latin_sync_t.ex
# packs it in a LATIN_PICK_BITS-wide field per (letter, case), so LATIN_PICK_MAX is
# the most that can ever be addressed -- keep it in step with state.h.  (It was a
# nibble/16 until the 6-bit widening.)
MAX_SLOTS    = 64                # LATIN_PICK_MAX in state.h (6-bit field)
FIRST_COL    = 2                 # column B


def _colname(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


HELPER_MIN   = _colname(FIRST_COL + MAX_SLOTS)       # first column past every slot
HELPER_MAX   = _colname(FIRST_COL + MAX_SLOTS + 1)
USAGE = """usage: _add_latin_variation.py [--dry-run] [--allow-asymmetric]
                              <xlsx> <letter> <char> [<letter> <char> ...]

  --dry-run           report the plan and change nothing
  --allow-asymmetric  permit adding to one case without its counterpart

e.g.  _add_latin_variation.py lang_lut.xlsx S 'S,' s 's,'   (real chars, not commas)"""


def die(msg):
    sys.exit(f"{os.path.basename(sys.argv[0])}: {msg}\n\n{USAGE}")


argv, DRY, ASYM = sys.argv[1:], False, False
while argv and argv[0].startswith("--"):
    if argv[0] == "--dry-run":
        DRY = True
    elif argv[0] == "--allow-asymmetric":
        ASYM = True
    else:
        die(f"unknown option {argv[0]!r}")
    argv = argv[1:]
if len(argv) < 3:
    die("expected an .xlsx and at least one <letter> <char> pair")
if len(argv) % 2 == 0:
    die("every <letter> needs exactly one <char> -- got an odd argument out")
XLSX, pairs = argv[0], list(zip(argv[1::2], argv[2::2]))
if not os.path.isfile(XLSX):
    die(f"file not found: {XLSX}")


def colname(n):                                  # 1-based index -> A1 letters
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def colidx(name):
    n = 0
    for ch in name:
        n = n * 26 + (ord(ch) - 64)
    return n


def cells_of(body):
    return re.findall(r'<c r="[A-Z]+\d+"[^>]*?(?:/>|>.*?</c>)', body, re.S)


def ref_of(cell):
    return re.search(r'r="([A-Z]+\d+)"', cell).group(1)


def col_of(ref):
    return colidx(re.match(r'([A-Z]+)', ref).group(1))


def row_xml(xml, row):
    m = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % row, xml, re.S)
    if not m:
        sys.exit(f"row {row} not found in {SHEET}")
    return m


def splice(xml, row, new_cells):
    """Insert cells (dict col -> xml) into <row r="row">, keeping column order."""
    m = row_xml(xml, row)
    out, pending = [], dict(new_cells)
    for c in cells_of(m.group(2)):
        ci = col_of(ref_of(c))
        for k in sorted(k for k in pending if k < ci):
            out.append(pending.pop(k))
        out.append(pending.pop(ci) if ci in pending else c)
    out.extend(pending[k] for k in sorted(pending))
    return xml[:m.start()] + m.group(1) + "".join(out) + m.group(3) + xml[m.end():]


def style_at(xml, row, target):
    """Reuse the neighbouring cell's style so the new cells match the rest."""
    for c in cells_of(row_xml(xml, row).group(2)):
        if ref_of(c) == target:
            m = re.search(r'\ss="(\d+)"', c)
            return f' s="{m.group(1)}"' if m else ""
    return ""


# --- creating a row for a letter that has no variations yet -------------------
#
# Only letters that already have at least one variation get a row, so the sheet
# carries 19 of 26 (B F M P Q V X are absent).  The cog fills the rest with NULL
# rows, which is why this was invisible until Irish needed Ḃ Ḟ Ṁ Ṗ.
#
# Two constraints decide where a new row can go:
#   * The cog walks ODD rows from 1 and STOPS at the first empty column A, so new
#     letter rows have to continue that block with no gap.  Order within it does
#     not matter -- the cog keys a dict by the letter label and emits A-Z then
#     a-z regardless -- so appending is fine and avoids renumbering the ~40 rows
#     (and their self-referencing DEC2HEX formulas) that inserting alphabetically
#     would touch.
#   * Rows 79/80 sit just past the block holding sheet-wide S/T aggregates with
#     an empty column A -- exactly what stops the scan.  They are human-facing
#     only (the cog never reads S/T), but they are in the way, so they move down
#     and their ranges are widened to cover the rows being added.

def block_end(xml):
    """Last odd row whose column A is filled — the end of the letter block."""
    last = 0
    for m in re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', xml, re.S):
        r = int(m.group(1))
        if r % 2 and any(ref_of(c).startswith("A") and col_of(ref_of(c)) == 1
                         for c in cells_of(m.group(2))):
            last = max(last, r)
    return last


def shift_trailing_rows(xml, first, delta):
    """Move every row >= `first` down by `delta`, rewriting refs to them.

    Deliberately narrow: it understands the two aggregate rows that actually
    live there and refuses anything else rather than risk mangling a formula it
    does not recognise.  A corrupted workbook would only surface later, in a
    generated table nobody diffed.
    """
    moved = []
    for m in list(re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', xml, re.S)):
        r = int(m.group(1))
        if r < first:
            continue
        for c in cells_of(m.group(2)):
            col = re.match(r'([A-Z]+)', ref_of(c)).group(1)
            if col not in (HELPER_MIN, HELPER_MAX, "Q"):
                sys.exit(f"refusing to create rows: unexpected cell {ref_of(c)} "
                         f"in trailing row {r} — this tool only knows how to "
                         f"relocate the sheet-wide S/T/Q aggregate rows")
        moved.append(r)
    for r in sorted(moved, reverse=True):          # high to low: refs stay unique
        m = row_xml(xml, r)
        body = m.group(2)
        body = re.sub(r'(<c r="([A-Z]+))%d"' % r, lambda mm: f'{mm.group(1)}{r + delta}"', body)
        # a formula referencing another moved row (row 80's DEC2HEX(S79))
        for other in moved:
            body = re.sub(r'\b([A-Z]+)%d\b' % other,
                          lambda mm, o=other: f"{mm.group(1)}{o + delta}", body)
        head = m.group(1).replace(f'r="{r}"', f'r="{r + delta}"', 1)
        xml = xml[:m.start()] + head + body + m.group(3) + xml[m.end():]
    # Widen the aggregate ranges (…BN2:BN92) so they span the newly added rows.
    # The block ends at the last letter row, so its DEC2HEX row -- and the end of
    # the range -- is `first - 1`.
    # ⚠️ Built per column rather than with a backreference: `\1` immediately
    # followed by the row number reads as group \192, which is a regex error, and
    # the character class was hardcoded to the old S/T columns besides.
    for col in (HELPER_MIN, HELPER_MAX):
        xml = re.sub(r'\b%s2:%s%d\b' % (col, col, first - 1),
                     f"{col}2:{col}{first - 1 + delta}", xml)
    return xml


def new_letter_rows(xml, letter, row, style_row):
    """Append the <row> pair for `letter`: the label row + its DEC2HEX row."""
    attrs = ('customFormat="false" ht="12.8" hidden="false" '
             'customHeight="false" outlineLevel="0" collapsed="false"')
    a_style = style_at(xml, style_row, f"A{style_row}")
    letter_row = (f'<row r="{row}" {attrs}><c r="A{row}"{a_style} t="inlineStr">'
                  f'<is><t xml:space="preserve">{escape(letter)}</t></is></c></row>')
    hex_row = f'<row r="{row + 1}" {attrs}></row>'
    return xml.replace("</sheetData>", letter_row + hex_row + "</sheetData>", 1)


# --- plan: locate each letter's row + its slot (reading is safe, saving is not)
from openpyxl import load_workbook
wb = load_workbook(XLSX, data_only=True)
sh = wb["latin_sup_ex"]
plan, taken = [], {}          # taken: row -> next free column, across this batch
# End of the contiguous letter block; rows for absent letters are appended here.
BLOCK_END = max((r for r in range(1, sh.max_row + 1, 2) if sh[f"A{r}"].value),
                default=0)
created, next_new = {}, BLOCK_END + 2
for letter, char in pairs:
    if len(char) != 1:
        die(f"{letter!r}: expected a single character, got {char!r}")
    if len(letter) != 1 or not letter.isascii() or not letter.isalpha():
        die(f"{letter!r}: expected a single ASCII letter")
    row = next((r for r in range(1, sh.max_row + 1, 2)
                if sh[f"A{r}"].value == letter), None)
    if row is None:
        # A letter with no variations yet has no row at all (B F M P Q V X).
        # Appending one is safe: the cog keys a dict by the label, so the sheet's
        # row order does not reach the generated table.
        row = created.get(letter)
        if row is None:
            row = created[letter] = next_new
            next_new += 2
    # ⚠️ Start from what THIS BATCH has already claimed, not from the worksheet:
    # two pairs naming the same letter both read the same original free column,
    # and the second splice then replaced the first -- silently dropping a
    # variation the tool reported as written.
    col = taken.get(row, FIRST_COL)
    while sh.cell(row=row, column=col).value is not None:
        if sh.cell(row=row, column=col).value == char:
            die(f"{letter}: {char!r} is already slot {col - FIRST_COL}")
        col += 1
    if col - FIRST_COL >= MAX_SLOTS:
        die(f"{letter}: row is full — {MAX_SLOTS} variations is the ceiling "
            f"(the picker stores its pick in a nibble; see MAX_SLOTS)")
    if any(p for p in plan if p[2] == row and p[1] == char):
        die(f"{letter}: {char!r} named twice in one command")
    taken[row] = col + 1
    plan.append((letter, char, row, col, f"{ord(char):X}"))   # DEC2HEX(UNICODE())
wb.close()

# The table is case-symmetric -- every variation exists in both cases -- and a
# half-pair is the shape of the mislabelled-row bug that shipped for ~2 years.
lonely = [l for l, _ in pairs
          if (l.swapcase() not in [x for x, _ in pairs])]
if lonely and not ASYM:
    die(f"one-case addition for {', '.join(sorted(set(lonely)))} -- the table is "
        f"case-symmetric, so pass both cases (or --allow-asymmetric if deliberate)")

for letter, char, row, col, hexv in plan:
    print(f"{letter}: append {char!r} U+{ord(char):04X} "
          f"({unicodedata.name(char, '?')}) at {colname(col)}{row} -> slot {col - 2}"
          + ("   [new row]" if letter in created else ""))
for letter, row in sorted(created.items(), key=lambda kv: kv[1]):
    print(f"{letter}: creating row {row} (letter had no variations before)")

if DRY:
    sys.exit(0)

with zipfile.ZipFile(XLSX) as z:
    xml = z.read(SHEET).decode("utf-8")     # read the one member; never extractall

if created:
    # Push the trailing sheet-wide aggregate rows clear of the letter block, then
    # append the new label rows so the cog's odd-row scan runs straight into them.
    xml = shift_trailing_rows(xml, BLOCK_END + 2, 2 * len(created))
    for letter, row in sorted(created.items(), key=lambda kv: kv[1]):
        xml = new_letter_rows(xml, letter, row, BLOCK_END)

for letter, char, row, col, hexv in plan:
    ref, hex_ref = f"{colname(col)}{row}", f"{colname(col)}{row + 1}"
    prev, prev_hex = f"{colname(col - 1)}{row}", f"{colname(col - 1)}{row + 1}"

    xml = splice(xml, row,
                 {col: f'<c r="{ref}"{style_at(xml, row, prev)} t="inlineStr">'
                       f'<is><t xml:space="preserve">{escape(char)}</t></is></c>'})
    xml = splice(xml, row + 1,
                 {col: f'<c r="{hex_ref}"{style_at(xml, row + 1, prev_hex)} t="str">'
                       f'<f aca="false">DEC2HEX(_xlfn.UNICODE({ref}))</f>'
                       f'<v>{hexv}</v></c>'})

    # Helper columns S (min) / T (max) are human-facing only -- cog never reads
    # them -- but a cache that disagrees with the row is worse than none.  S's
    # range is maintained to span exactly the filled slots, so extend it; T
    # already spans past the end, so only its cached value moves.
    m = row_xml(xml, row + 1)
    fixed = []
    for c in cells_of(m.group(2)):
        r = ref_of(c)
        # ⚠️ Not every helper cell carries a cached aggregate: some rows hold an
        # EMPTY placeholder that has a style and nothing else (`<c r="S11" s="12"/>`,
        # on the d/f/h/j/l rows).  Reading a <v> out of one crashed the tool with a
        # bare AttributeError -- latent until the first addition to such a letter,
        # which is what Irish's Ḋ/ḋ finally hit.  There is no cached value to keep
        # in step there, so leave the placeholder alone.
        cached = re.search(r'<v>(-?\d+)</v>', c)
        if r.startswith(HELPER_MIN) and cached:
            c = re.sub(r'(<f[^>]*>MIN\()(.*?)(\)</f>)',
                       lambda mm: f"{mm.group(1)}{mm.group(2)},HEX2DEC({hex_ref}){mm.group(3)}",
                       c)
            c = re.sub(r'<v>-?\d+</v>', f'<v>{min(int(cached.group(1)), ord(char))}</v>', c)
        elif r.startswith(HELPER_MAX) and cached:
            c = re.sub(r'<v>-?\d+</v>', f'<v>{max(int(cached.group(1)), ord(char))}</v>', c)
        fixed.append(c)
    xml = xml[:m.start()] + m.group(1) + "".join(fixed) + m.group(3) + xml[m.end():]

# The splices above are string surgery -- deliberately, because anything that
# re-serialises the sheet would rewrite bytes we are trying to leave alone.  So
# prove the result still parses before it replaces a working workbook: a bad
# regex would otherwise ship a corrupt .xlsx that only fails when someone opens it.
try:
    minidom.parseString(xml.encode("utf-8"))
except Exception as e:
    sys.exit(f"refusing to write: the patched {SHEET} is not well-formed XML ({e})")

out = XLSX + ".new"
with zipfile.ZipFile(XLSX) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
    for it in zin.infolist():
        zout.writestr(it, xml.encode("utf-8") if it.filename == SHEET
                          else zin.read(it.filename))
shutil.move(out, XLSX)
print(f"wrote {XLSX} ({len(plan)} variation(s))")
