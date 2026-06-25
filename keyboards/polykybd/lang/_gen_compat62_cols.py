#!/usr/bin/env python3
"""Generate key_lut columns for the 2026-06 "compat easy-win" batch — fold/clone
languages that need NO new font (the host compatibility map switches the OS
layout). Ranked by computer/internet users, 4-15 per region tab.

  Americas (15) Spanish   -> clone es-MX (latam)
  Europe   (8)  mixed     -> clone de-DE/fr-BE/es-ES/hr-HR/de-CH/da-DK, fold en-IE
  M.East   (10) Arabic    -> clone ar-SA (ara)
  Africa   (15) ar/fr/pt  -> clone ar-SA/fr-FR/pt-PT, fold en-*/sw-TZ
  Asia     (8)  mixed     -> clone bn-IN/ru-RU, fold en-*
  Oceania  (6)  mixed     -> clone fr-FR/sm-WS, fold en-*

Every distinct-layout entry is a CLONE (copies all four variants incl. AltGr),
so AltGr legends are inherited from the source layout. Only true US-QWERTY
locales (English variants, Swahili-TZ) become empty folds (en-US fallback caps).

No new named glyphs (all clones reuse existing glyphs; folds are empty).

Outputs: /tmp/compat62_cols.json (for _patch_xlsx.py)
"""
import json
from openpyxl import load_workbook

XLSX = "lang_lut.xlsx"

S = lambda v: ["str", v]
N = lambda v: ["num", v]

# id-ID-style fold settings (same block en-AU/en-NZ/... use).
SET_FOLD = {56: [None, S("HIDE"), None, None], 57: [None, S("HIDE"), None, None],
            58: [None, N(35), None, None],     59: [None, N(0), None, None],
            60: [None, N(35), None, None],     61: [None, N(0), None, None]}

def clone_col(sheet, langs, src):
    base = langs[src]
    out = {}
    for r in range(2, 62):                       # rows 2..61 incl. settings (56-61)
        cells = []
        for k in range(4):
            v = sheet.cell(row=r, column=base + k).value
            if v is None:
                cells.append(None)
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                cells.append(N(int(v)))
            else:
                cells.append(S(str(v)))
        if any(c is not None for c in cells):
            out[r] = cells
    return out

wb = load_workbook(XLSX, data_only=True, read_only=True)
sh = wb["key_lut"]
langs = {}
i = 0
while True:
    v = sh.cell(row=1, column=2 + i * 4).value
    if not v:
        break
    langs[v] = 2 + i * 4
    i += 1

# ── batch spec: name -> ("fold" | source-column-name) ────────────────────────
FOLD = "fold"
SPECMAP = [
    # Americas (15) — Spanish, clone es-MX (latam)
    ("es-CO", "es-MX"), ("es-PE", "es-MX"), ("es-VE", "es-MX"), ("es-CL", "es-MX"),
    ("es-EC", "es-MX"), ("es-GT", "es-MX"), ("es-DO", "es-MX"), ("es-BO", "es-MX"),
    ("es-PY", "es-MX"), ("es-CR", "es-MX"), ("es-SV", "es-MX"), ("es-HN", "es-MX"),
    ("es-PA", "es-MX"), ("es-UY", "es-MX"), ("es-NI", "es-MX"),
    # Europe (8)
    ("de-AT", "de-DE"), ("nl-BE", "fr-BE"), ("ca-ES", "es-ES"), ("en-IE", FOLD),
    ("bs-BA", "hr-HR"), ("fr-CH", "de-CH"), ("sl-SI", "hr-HR"), ("fo-FO", "da-DK"),
    # Middle East & Caucasus (10) — Arabic, clone ar-SA (ara)
    ("ar-AE", "ar-SA"), ("ar-SY", "ar-SA"), ("ar-JO", "ar-SA"), ("ar-LB", "ar-SA"),
    ("ar-YE", "ar-SA"), ("ar-KW", "ar-SA"), ("ar-OM", "ar-SA"), ("ar-PS", "ar-SA"),
    ("ar-QA", "ar-SA"), ("ar-BH", "ar-SA"),
    # Africa (15)
    ("ar-DZ", "ar-SA"), ("ar-SD", "ar-SA"), ("ar-TN", "ar-SA"), ("ar-LY", "ar-SA"),
    ("fr-CD", "fr-FR"), ("fr-CI", "fr-FR"), ("fr-CM", "fr-FR"), ("fr-SN", "fr-FR"),
    ("fr-MG", "fr-FR"), ("en-GH", FOLD), ("en-UG", FOLD), ("en-ZM", FOLD),
    ("sw-TZ", FOLD), ("pt-AO", "pt-PT"), ("pt-MZ", "pt-PT"),
    # Asia (8)
    ("bn-BD", "bn-IN"), ("en-IN", FOLD), ("en-PK", FOLD), ("en-PH", FOLD),
    ("en-SG", FOLD), ("en-LK", FOLD), ("ky-KG", "ru-RU"), ("tg-TJ", "ru-RU"),
    # Oceania (6)
    ("en-GU", FOLD), ("en-SB", FOLD), ("en-VU", FOLD), ("en-FM", FOLD),
    ("fr-NC", "fr-FR"), ("to-TO", "sm-WS"),
]

SPEC = {}
for name, src in SPECMAP:
    if src == FOLD:
        SPEC[name] = dict(SET_FOLD)
    else:
        assert src in langs, f"source column {src} not found"
        SPEC[name] = clone_col(sh, langs, src)

ORDER = [n for n, _ in SPECMAP]
assert len(ORDER) == len(set(ORDER)) == 62, f"expected 62 unique, got {len(ORDER)}"

spec_json = {n: {str(r): v for r, v in cols.items()} for n, cols in SPEC.items()}
json.dump(spec_json, open("/tmp/compat62_cols.json", "w"), indent=1)
print("order:", ",".join(ORDER))
print(f"{len(ORDER)} languages "
      f"({sum(1 for _,s in SPECMAP if s==FOLD)} folds, "
      f"{sum(1 for _,s in SPECMAP if s!=FOLD)} clones)")
