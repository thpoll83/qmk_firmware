#!/usr/bin/env python3
"""Regenerate the REGION_OFFSET / REGION_LANGS tables in lang_layer.c from the
single source of truth (lang_lut.xlsx language order) + the host region map
(PolyKybdHost/polyhost/services/lang_regions.py). Mirrors the host's stable
grouping: region order = LANG_REGION_ORDER; within a region, sort by country
code then enum (= GET_LANG_LIST) order. Prints the C block to stdout.
"""
import sys, os
from collections import defaultdict
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, "/home/user/PolyKybdHost/polyhost/services")
from lang_regions import LANG_REGION, LANG_REGION_ORDER, LANG_REGION_OVERRIDE

sh = load_workbook(os.path.join(HERE, "lang_lut.xlsx"), data_only=True, read_only=True)["key_lut"]
langs, i = [], 0
while True:
    v = sh.cell(row=1, column=2 + i * 4).value
    if not v:
        break
    langs.append(v)
    i += 1

def region_of(code):
    if code.replace("-", "") in LANG_REGION_OVERRIDE:
        return LANG_REGION_OVERRIDE[code.replace("-", "")]
    return LANG_REGION.get(code.split("-")[1].upper(), "Other")

buckets = defaultdict(list)
for idx, code in enumerate(langs):
    buckets[region_of(code)].append((code.split("-")[1].upper(), idx, code))

SHORT = {"Americas": "Americas", "Europe": "Europe",
         "Middle East & Caucasus": "Middle East & Caucasus", "Africa": "Africa",
         "Asia": "Asia", "Oceania": "Oceania"}

order, offsets, comments = [], [0], []
for region in LANG_REGION_ORDER:
    items = sorted(buckets[region], key=lambda t: (t[0], t[1]))
    codes = [t[2] for t in items]
    idxs = [t[1] for t in items]
    order += idxs
    offsets.append(len(order))
    # wrap the code list as comment lines
    head = f"    // {SHORT[region]} ({len(items)}): "
    pad = "    //" + " " * (len(head) - len("    //"))
    line, lines = head, []
    for j, c in enumerate(codes):
        tok = c + (" " if j < len(codes) - 1 else "")
        if len(line) + len(tok) > 96:
            lines.append(line.rstrip())
            line = pad + tok
        else:
            line += tok
    lines.append(line.rstrip())
    comments.append("\n".join(lines))

# emit C
print(f"static const uint8_t REGION_OFFSET[NUM_LANG_REGIONS + 1] = {{")
print("    " + ", ".join(str(o) for o in offsets) + ",")
print("};\n")
print("static const uint8_t REGION_LANGS[NUM_LANG] = {")
pos = 0
for r, region in enumerate(LANG_REGION_ORDER):
    print(comments[r])
    seg = order[offsets[r]:offsets[r + 1]]
    line = "    "
    for v in seg:
        tok = f"{v:>3}, "
        if len(line) + len(tok) > 84:
            print(line.rstrip())
            line = "    "
        line += tok
    print(line.rstrip())
print("};")
print(f"// total = {len(order)} (NUM_LANG)", file=sys.stderr)
print(f"// offsets = {offsets}", file=sys.stderr)
