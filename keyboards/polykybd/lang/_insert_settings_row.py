#!/usr/bin/env python3
"""Insert a new SETTINGS row into the key_lut sheet (sheet2.xml) of lang_lut.xlsx.

The settings block is the run of rows whose column A holds a ``{category.name}``
key (today rows 56-61). cog collects it by walking column A until the first empty
cell, so a new setting has to be INSERTED inside that run -- appending below the
`lower/upper/caps/ALT Gr` legend row that follows it would never be seen.

Inserting means renumbering every row below, which is why this is a script and not
a hand edit.

⚠️ Row numbers live in FOUR places, not one, and only two of them are inside
``<sheetData>``. ``<row r>`` and ``<c r>`` are; ``<mergeCell ref>`` and
``<hyperlink ref>`` sit AFTER it, so a renumbering pass that slices the tail of
sheetData never reaches them. This sheet has 27 merges and 10 hyperlinks anchored on
the ``lower/upper/caps/ALT Gr`` legend row -- the row an insert pushes down -- so
missing them leaves those refs pointing one row short.

That failure is SILENT and asymmetric between openpyxl's two readers, which is what
makes it nasty: ``read_only=True`` streams row elements and reports the sheet as
correct, while the eager loader keys on cell refs, materialises a phantom cell where
the stale hyperlink still points, and hands back its ``display`` text as that cell's
value. cog uses the eager loader, so the first inserted settings row came out with
``https://www.branah.com/english`` as its en-US value (2026-09-02). Verify an edit
with BOTH readers, not one -- ``_verify_readers_agree()`` below now does it for you.

An earlier version of this docstring claimed the workbook has "no merged cells
outside row 1". It never did.

The new row is CLONED from the last settings row, so it inherits that row's cell
styles and column layout exactly; only the values differ.

⚠️ openpyxl's EAGER loader mis-reads a couple of shared-string cells in the legend
row (`upper`/`caps` come back as None) while `read_only=True`, the raw XML and cog all
read them correctly. This note used to say "after an insert"; it is not -- the same
disagreement is there in the UNTOUCHED workbook, so an insert neither causes nor
worsens it. So do not chase that; check a workbook edit with
`load_workbook(..., read_only=True)`, and judge the result by whether the cog
regeneration is purely ADDITIVE.

usage: _insert_settings_row.py <xlsx> <{key.name}> <variation 0-3> <lang,lang,...> [value]

  variation  0 lower / 1 upper / 2 caps / 3 ALT Gr -- which of the language's four
             sub-columns carries the value (VAR_SMALL/SHIFT/CAPS/ALTGR in the C).
  value      what to write for each named language (default 1); every other
             language is left blank, which cog emits as 0. An EMPTY lang list adds
             the row with no language set, which is how an opt-in nothing uses yet
             still becomes selectable in the keycap tuner.
"""
import os
import re
import shutil
import sys
import tempfile
import zipfile
from xml.sax.saxutils import escape

def _verify_readers_agree(path, new_row, span=12):
    """Both openpyxl readers must see the same settings block, or the edit is bad.

    The streaming reader walks row ELEMENTS; the eager one keys on cell REFS. A row
    number left behind in a `<mergeCell>` or `<hyperlink>` shows up only in the
    second, as a phantom cell carrying the link's display text -- and cog uses the
    eager loader. Comparing the two is the cheapest possible check for that whole
    class, so it runs on every insert rather than being left to whoever remembers.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:                      # not a reason to fail the edit
        print("note: openpyxl unavailable, skipped the two-reader check", file=sys.stderr)
        return
    lo, hi = max(1, new_row - span), new_row + span
    wb = load_workbook(path, read_only=True, data_only=True)
    sh = wb["key_lut"]
    stream = {i: r[:6] for i, r in
              enumerate(sh.iter_rows(min_row=lo, max_row=hi, max_col=6, values_only=True), start=lo)}
    wb.close()
    wb = load_workbook(path, data_only=True)
    sh = wb["key_lut"]
    eager = {i: tuple(sh.cell(row=i, column=c).value for c in range(1, 7)) for i in range(lo, hi + 1)}
    wb.close()
    # ⚠️ Compare ASYMMETRICALLY, and the direction is the whole point.
    #   eager has a value streaming does NOT  -> a phantom cell conjured by a stale
    #                                            ref. This is the corruption.
    #   eager has None where streaming has a value -> the pre-existing shared-string
    #                                            quirk (see the docstring); benign.
    # A symmetric compare fires on the legend row of an UNTOUCHED workbook, so it
    # would be a guard nobody could keep green.
    bad = []
    for i, cells in eager.items():
        ref = stream.get(i) or ()
        for j, v in enumerate(cells):
            if v is not None and v != (ref[j] if j < len(ref) else None):
                bad.append(f"row {i} col {j + 1}: eager={v!r} stream="
                           f"{ref[j] if j < len(ref) else None!r}")
    if bad:
        sys.exit("the eager openpyxl reader sees values the streaming one does not, at "
                 + "; ".join(bad[:4])
                 + " -- a row number was left behind outside <sheetData>. The file has "
                   "been written and is WRONG; restore it. See the docstring.")


if len(sys.argv) < 5:
    sys.exit(__doc__)
XLSX, KEY, VARIATION = sys.argv[1], sys.argv[2], int(sys.argv[3])
# An EMPTY list is legitimate: a row every layout opts out of by default still has to
# exist, because the tuner can only offer a setting the spreadsheet already carries.
LANGS = [s for s in sys.argv[4].split(",") if s]
VALUE = sys.argv[5] if len(sys.argv) > 5 else "1"
# The cell is written with t="n" below, so a non-numeric value would produce a
# numeric cell holding text -- a workbook Excel and openpyxl both reject, from a
# script whose whole job is not to corrupt this file.
try:
    int(VALUE)
except ValueError:
    sys.exit(f"value must be an integer: {VALUE!r}")
if not (0 <= VARIATION <= 3):
    sys.exit("variation must be 0..3")
if not os.path.isfile(XLSX):
    sys.exit(f"file not found: {XLSX}")

tmp = tempfile.mkdtemp(prefix="insrow_")
try:
    with zipfile.ZipFile(XLSX) as z:
        z.extractall(tmp)
    sheet = os.path.join(tmp, "xl/worksheets/sheet2.xml")
    xml = open(sheet, encoding="utf-8").read()

    # Anything else that carries a row number and is NOT renumbered below. Fail loudly
    # rather than silently shifting the sheet's meaning -- that is the whole lesson of
    # the hyperlink/mergeCell miss described in the docstring.
    for bad, what in ((r"<f>", "formulas"),
                      (r"<conditionalFormatting", "conditional formatting"),
                      (r"<dataValidation", "data validation"),
                      (r"<autoFilter", "an autofilter"),
                      (r"<ignoredErrors", "ignored-error ranges")):
        if re.search(bad, xml):
            sys.exit(f"sheet2 has {what}; renumbering would break it -- see the docstring")

    shared = open(os.path.join(tmp, "xl/sharedStrings.xml"), encoding="utf-8").read()
    strings = ["".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S))
               for si in re.findall(r"<si>(.*?)</si>", shared, re.S)]

    def col_index(name):
        n = 0
        for ch in name:
            n = n * 26 + (ord(ch) - 64)
        return n

    def col_name(n):
        s = ""
        while n:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    rows = {int(r): (m.start(), m.end()) for m, r in
            ((m, m.group(1)) for m in re.finditer(r'<row r="(\d+)".*?</row>', xml, re.S))}

    def cell_text(row, col):
        """The displayed text of one cell, resolving a shared-string reference."""
        seg = xml[rows[row][0]:rows[row][1]]
        m = re.search(r'<c r="%s%d"[^>]*?(?:/>|>(.*?)</c>)' % (col_name(col), row), seg, re.S)
        if not m or not m.group(1):
            return None
        body = m.group(1)
        if 't="s"' in m.group(0):
            return strings[int(re.search(r"<v>(\d+)</v>", body).group(1))]
        v = re.search(r"<v>(.*?)</v>", body, re.S)
        it = re.search(r"<t[^>]*>(.*?)</t>", body, re.S)
        return (v or it).group(1) if (v or it) else None

    # The settings block: the contiguous run of {key} rows in column A.
    settings = sorted(r for r in rows if (cell_text(r, 1) or "").startswith("{"))
    if not settings or settings != list(range(settings[0], settings[-1] + 1)):
        sys.exit(f"settings rows are not contiguous: {settings}")
    last, new_row = settings[-1], settings[-1] + 1
    if any((cell_text(r, 1) or "") == KEY for r in settings):
        sys.exit(f"{KEY} already exists")

    langs = []                                   # header row 1, one name per 4 columns
    c = 2
    while (name := cell_text(1, c)):
        langs.append(name)
        c += 4
    unknown = [lang for lang in LANGS if lang not in langs]
    if unknown:
        sys.exit(f"unknown language(s): {unknown}")
    targets = {2 + 4 * langs.index(lang) + VARIATION for lang in LANGS}

    # Clone the last settings row: same row attributes, same per-cell styles.
    seg = xml[rows[last][0]:rows[last][1]]
    head = re.match(r"<row [^>]*>", seg).group(0)
    out = [head.replace('r="%d"' % last, 'r="%d"' % new_row, 1)]
    for c in re.findall(r'<c r="[A-Z]+\d+"[^>]*?(?:/>|>.*?</c>)', seg, re.S):
        ref = re.search(r'r="([A-Z]+)(\d+)"', c)
        col, style = col_index(ref.group(1)), re.search(r'\ss="\d+"', c)
        style = style.group(0) if style else ""
        r = f'{col_name(col)}{new_row}'
        if col == 1:
            out.append(f'<c r="{r}"{style} t="inlineStr"><is><t xml:space="preserve">{escape(KEY)}</t></is></c>')
        elif col in targets:
            out.append(f'<c r="{r}"{style} t="n"><v>{escape(VALUE)}</v></c>')
        else:
            out.append(f'<c r="{r}"{style}/>')
    out.append("</row>")
    new_xml_row = "".join(out)

    # Renumber every row at or below the insertion point, then splice the new one in.
    def bump(m):
        n = int(m.group(2))
        return m.group(1) + (str(n + 1) if n >= new_row else str(n)) + m.group(3)

    tail_start = rows[new_row][0] if new_row in rows else rows[last][1]
    head_xml, tail_xml = xml[:tail_start], xml[tail_start:]
    tail_xml = re.sub(r'(<row r=")(\d+)(")', bump, tail_xml)
    tail_xml = re.sub(r'(<c r="[A-Z]+)(\d+)(")', bump, tail_xml)
    xml = head_xml + new_xml_row + tail_xml

    # ⚠️ mergeCells and hyperlinks live AFTER </sheetData>, so the tail slice above
    # never touched them. Renumber across the WHOLE document -- these patterns cannot
    # match a <row>/<c> element, so nothing gets bumped twice.
    xml = re.sub(r'(<hyperlink ref="[A-Z]+)(\d+)(")', bump, xml)

    def bump_merge(m):
        return (m.group(1) + str(int(m.group(2)) + 1 if int(m.group(2)) >= new_row else int(m.group(2)))
                + m.group(3) + str(int(m.group(4)) + 1 if int(m.group(4)) >= new_row else int(m.group(4)))
                + m.group(5))

    xml = re.sub(r'(<mergeCell ref="[A-Z]+)(\d+)(:[A-Z]+)(\d+)(")', bump_merge, xml)

    m = re.search(r'<dimension ref="([A-Z]+)(\d+):([A-Z]+)(\d+)"/>', xml)
    if m:
        xml = xml.replace(m.group(0),
                          f'<dimension ref="{m.group(1)}{m.group(2)}:{m.group(3)}{int(m.group(4)) + 1}"/>')

    open(sheet, "w", encoding="utf-8").write(xml)
    out_xlsx = XLSX + ".new"
    with zipfile.ZipFile(XLSX) as zin, zipfile.ZipFile(out_xlsx, "w", zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            data = open(sheet, "rb").read() if it.filename == "xl/worksheets/sheet2.xml" else zin.read(it.filename)
            zout.writestr(it, data)
    shutil.move(out_xlsx, XLSX)
    _verify_readers_agree(XLSX, new_row)
    print(f"inserted {KEY} at row {new_row}; set {len(targets)} language(s) to {VALUE}")
finally:
    shutil.rmtree(tmp, ignore_errors=True)
